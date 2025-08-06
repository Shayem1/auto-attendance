from tkinter import *
import tkinter as tk
import customtkinter
import xlsxwriter
import time
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

def open(file):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file)

    # Select the active sheet
    sheet = wb.active

    # Create an empty list to store the data from the first column
    students = []

    # Loop through the rows and extract data from the first column (column A)
    for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):  # Skipping the header row
        students.append(row[0]) 
    wb.close()

    return students


#button timeouts
def active():
    global first_button
    first_button.configure(state=NORMAL)


def phase_3(students):
    global driver, first_button, label

    label.configure(text="Go to Page 4 and press continue and add to\nmerge documents, once completed, repeat\nfor all pages", font=("arial",20))

    data_list = []
    checkbox = []
    index_list = []

    found_data = driver.find_elements(By.XPATH ,'//*[@id="contactsTable"]/tbody/tr/td[4]')
    checkbox_data = driver.find_elements(By.XPATH ,'//*[@id="contactsTable"]/tbody/tr/td[1]/input[1]')

    for i in found_data:
        data_list.append(i.text)

    for index, value in enumerate(data_list):
        if value in students:
            index_list.append(index)

    for i in checkbox_data:
            checkbox.append(i)

    for i in index_list:
        checkbox[i].click()


def phase_2(link):

    #move to
    global status, driver, label, first_button
    status = 2

    #opens the next webpage
    driver.get(link)

    #next instructions
    label.configure(text="Please select optional id\nand sort by optional id", font=("arial",20))

    if link == "https://ibma.app.axcelerate.com/management/management2/ContactsSearch.cfm":
        first_button.configure(text="Continue", font=("arial",30), corner_radius=30, state=DISABLED,  command= lambda: phase_3(students_IBMA))
    elif link == "https://ibmnsw.app.axcelerate.com/management/management2/ContactsSearch.cfm":
        first_button.configure(text="Continue", font=("arial",30), corner_radius=30, state=DISABLED,  command= lambda: phase_3(students_IBMNSW))
    

    #ensures user does the action
    GUI.after(5000, active)

def phase_1(link):
    #move to first stage
    global status, label, first_button, driver, second_button
    second_button.destroy()
    status = 1

    #opens chrome page
    driver = webdriver.Chrome()
    driver.get(link)
    

    #adds instructions for the next stage
    label.configure(text="Please Login to any account\n with access to merge mailout groups\n\nPress continue AFTER you login")
    first_button.configure(state=DISABLED, text="Continue", command= lambda: phase_2(f"{link}management2/ContactsSearch.cfm"))
    

    GUI.after(20000, active)

#helps code remember what phase it is on
status = 0

#Gui setup
GUI = customtkinter.CTk()
GUI.title("Tasker")
GUI.geometry("450x200")
GUI.minsize(width=450, height=300)
GUI.maxsize(width=450, height=300)
customtkinter.set_appearance_mode("dark")

#setting up info text for phase 0
label = customtkinter.CTkLabel(master=GUI, text="Hello, Tasker will guide you on how\nto automatically make merge out groups\nin seconds from the attendance sheet", font=("arial",20))
label.pack(pady=15)

first_button = customtkinter.CTkButton(master = GUI, text="IBMA", font=("arial",30), corner_radius=30, command= lambda: phase_1("https://ibma.app.axcelerate.com/management/"))
first_button.pack(pady=30)

second_button = customtkinter.CTkButton(master = GUI, text="IBMNSW", font=("arial",30), corner_radius=30, command= lambda: phase_1("https://ibmnsw.app.axcelerate.com/management/"))
second_button.pack(pady=30)

students_IBMA = open("IBMA.xlsx")
students_IBMNSW = open("IBMNSW.xlsx")

GUI.mainloop()