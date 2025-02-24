import os
import openpyxl
import xlsxwriter
from datetime import datetime

# Function to find the most recent .xlsx file in the directory
def get_most_recent_file(directory):
    files = [f for f in os.listdir(directory) if f.endswith('.xlsx')]
    if not files:
        return None
    latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(directory, f)))
    return os.path.join(directory, latest_file)

# Function to process the data
def process_excel_file(file_path):
    # Load the workbook and sheet
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    # Create two new workbooks using xlsxwriter
    ibmnsw_wb = xlsxwriter.Workbook("IBMNSW.xlsx")
    ibmnsw_ws = ibmnsw_wb.add_worksheet()
    
    ibma_wb = xlsxwriter.Workbook("IBMA.xlsx")
    ibma_ws = ibma_wb.add_worksheet()

    # Define row counters for both workbooks
    ibmnsw_row = 0
    ibma_row = 0
    
    # Loop through rows in the active sheet
    for row in sheet.iter_rows(min_row=2):  # Assuming the first row is the header
        first_column = row[0].value
        eighth_column = row[7].value
        
        # Check for IBMNSW criteria
        if isinstance(first_column, str) and len(first_column) == 9 and first_column.startswith("IBM") and (eighth_column is None or str(eighth_column).strip() == ""):
            ibmnsw_ws.write_row(ibmnsw_row, 0, [cell.value for cell in row])
            ibmnsw_row += 1

        # Check for IBMA criteria
        if isinstance(first_column, str) and len(first_column) == 8 and first_column.isdigit() and (eighth_column is None or str(eighth_column).strip() == ""):
            ibma_ws.write_row(ibma_row, 0, [cell.value for cell in row])
            ibma_row += 1

    # Close the new workbooks
    ibmnsw_wb.close()
    ibma_wb.close()

# Main function
def main():
    # Set the directory to look for the most recent .xlsx file
    directory = os.getcwd()  # Current working directory
    recent_file = get_most_recent_file(directory)

    if recent_file:
        print(f"Processing the most recent file: {recent_file}")
        process_excel_file(recent_file)
        print("Processing complete. Files 'IBMNSW.xlsx' and 'IBMA.xlsx' have been created.")
    else:
        print("No .xlsx files found in the directory.")

if __name__ == "__main__":
    main()